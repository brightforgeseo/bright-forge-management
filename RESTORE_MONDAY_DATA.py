#!/usr/bin/env python3
"""
Parse Monday.com export and generate SQL to restore to Supabase
"""

import openpyxl
import json
import os
from datetime import datetime
from pathlib import Path

# Load updates/comments
updates_file = Path('/Users/user/Downloads/bright-forge-portal/monday_updates.json')
UPDATES = {}
if updates_file.exists():
    with open(updates_file, 'r') as f:
        UPDATES = json.load(f)

def clean_string(s):
    """Clean string for SQL"""
    if s is None:
        return ''
    s = str(s).replace("'", "''").replace('\n', ' ').replace('\r', ' ')
    return s[:1000]  # Limit length

def parse_date(date_val):
    """Parse date to YYYY-MM-DD format"""
    if isinstance(date_val, datetime):
        return date_val.strftime('%Y-%m-%d')
    if isinstance(date_val, str):
        try:
            dt = datetime.strptime(date_val, '%Y-%m-%d')
            return dt.strftime('%Y-%m-%d')
        except:
            pass
    return '2025-11-20'  # Default date

def parse_board_file(filepath):
    """Parse a Monday.com board export file - handles multiple formats"""
    wb = openpyxl.load_workbook(filepath)
    sheet = wb.active

    # Get board name from first cell
    board_name = str(sheet.cell(1, 1).value or Path(filepath).stem.split('_', 1)[1])

    # Detect format by checking row 2
    row2_val = str(sheet.cell(2, 1).value or '')

    # Format 1: Has group name in row 2 (older format with groups)
    # Format 2: Has "Onboarding" or other text in row 2 (newer format, single group)

    groups = []
    current_group = None
    current_group_tasks = []

    # Find header row - could be row 3 or row 5
    header_row_idx = None
    for row_idx in [3, 5]:
        row = [str(sheet.cell(row_idx, col).value or '') for col in range(1, 12)]
        if 'Name' in row[0] and any('Person' in str(cell) for cell in row):
            header_row_idx = row_idx
            break

    # Check if this is the simpler format (no group separators)
    is_simple_format = header_row_idx is not None

    if is_simple_format:
        # Single group format - use row 2 or row before headers as group name
        # Need to detect column layout from header row
        header_row = [str(sheet.cell(header_row_idx, col).value or '').lower() for col in range(1, 12)]

        # Find column indices
        name_col = 0
        person_col = next((i for i, h in enumerate(header_row) if 'person' in h), None)
        status_col = next((i for i, h in enumerate(header_row) if 'status' in h), None)
        date_col = next((i for i, h in enumerate(header_row) if 'date' in h), None)
        priority_col = next((i for i, h in enumerate(header_row) if 'priority' in h), None)
        client_col = next((i for i, h in enumerate(header_row) if 'client' in h or 'file' in h), None)
        item_col = next((i for i, h in enumerate(header_row) if 'item' in h and 'id' in h), None)

        # Get group name from row before headers
        group_row_idx = header_row_idx - 1
        group_name = str(sheet.cell(group_row_idx, 1).value or 'Tasks')
        if group_name == board_name or 'http' in group_name.lower():
            group_name = 'Tasks'

        current_group = clean_string(group_name)
        current_group_tasks = []

        # Start parsing from row after headers
        for row_idx in range(header_row_idx + 1, sheet.max_row + 1):
            # Get all columns
            row = [sheet.cell(row_idx, col).value for col in range(1, 12)]

            # Item ID indicates a valid task row
            item_id = row[item_col] if item_col and len(row) > item_col else None

            # Skip if no item ID or no task name
            if not item_id or not row[name_col]:
                continue

            # Get comments for this task from updates
            task_comments = UPDATES.get(str(item_id), [])

            task = {
                'title': clean_string(row[name_col]),
                'person': clean_string(row[person_col]) if person_col and len(row) > person_col else '',
                'status': clean_string(row[status_col]) if status_col and len(row) > status_col else '',
                'date': parse_date(row[date_col]) if date_col and len(row) > date_col else '2025-11-20',
                'priority': clean_string(row[priority_col]) if priority_col and len(row) > priority_col else '',
                'client_sheet': clean_string(row[client_col]) if client_col and len(row) > client_col else '',
                'worksheet': '',  # Not in this format
                'item_id': str(item_id),
                'comments': task_comments
            }
            current_group_tasks.append(task)

        if current_group_tasks:
            groups.append({
                'title': current_group,
                'tasks': current_group_tasks
            })

    else:
        # Original multi-group format
        row_idx = 4  # Start after board info

        while row_idx <= sheet.max_row:
            row = [sheet.cell(row_idx, col).value for col in range(1, 9)]

            # Check if this is a group header (no item ID in last column)
            if row[0] and not row[7] and row[0] != 'Name':
                # Save previous group if exists
                if current_group and current_group_tasks:
                    groups.append({
                        'title': current_group,
                        'tasks': current_group_tasks
                    })

                # Start new group
                current_group = clean_string(row[0])
                current_group_tasks = []
                row_idx += 2  # Skip header row
                continue

            # Check if this is a task row (has item ID)
            if row[7]:  # Item ID exists
                # Get comments for this task from updates
                task_comments = UPDATES.get(str(row[7]), [])

                task = {
                    'title': clean_string(row[0]),
                    'person': clean_string(row[1]),
                    'status': clean_string(row[2]),
                    'date': parse_date(row[3]),
                    'priority': clean_string(row[4]),
                    'client_sheet': clean_string(row[5]),
                    'worksheet': clean_string(row[6]),
                    'item_id': str(row[7]),
                    'comments': task_comments
                }
                current_group_tasks.append(task)

            row_idx += 1

        # Save last group
        if current_group and current_group_tasks:
            groups.append({
                'title': current_group,
                'tasks': current_group_tasks
            })

    return {
        'name': board_name,
        'groups': groups
    }

# Parse all board files
boards_dir = Path('~/Downloads/boards').expanduser()
all_boards = []

print("Parsing Monday.com boards...")
for filepath in sorted(boards_dir.glob('*.xlsx')):
    try:
        board_data = parse_board_file(filepath)
        if board_data['groups']:
            all_boards.append(board_data)
            print(f"✓ {board_data['name']}: {len(board_data['groups'])} groups, {sum(len(g['tasks']) for g in board_data['groups'])} tasks")
    except Exception as e:
        print(f"✗ Error parsing {filepath.name}: {e}")

print(f"\nTotal boards: {len(all_boards)}")
print(f"Total tasks: {sum(sum(len(g['tasks']) for g in b['groups']) for b in all_boards)}")

# Save to JSON for review
with open('/Users/user/Downloads/bright-forge-portal/monday_data.json', 'w') as f:
    json.dump(all_boards, f, indent=2)

print("\nSaved data to monday_data.json")
print(f"\nFirst board sample: {all_boards[0]['name']}")
print(f"Groups: {[g['title'] for g in all_boards[0]['groups'][:3]]}")
