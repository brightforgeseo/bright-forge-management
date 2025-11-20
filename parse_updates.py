#!/usr/bin/env python3
"""
Parse Monday.com updates (comments) and match them to tasks by Item ID
"""

import openpyxl
import json
from pathlib import Path
from datetime import datetime

def parse_updates_file(filepath):
    """Parse a Monday.com updates file"""
    wb = openpyxl.load_workbook(filepath)
    sheet = wb.active

    # Row 2 has the actual headers (Row 1 is board name)
    # Columns: Item ID, Item Name, Content Type, Content Type, User, Created At, Update Content, Likes Count, Asset IDs

    updates_by_item_id = {}

    for row_idx in range(3, sheet.max_row + 1):  # Start from row 3 (after headers)
        item_id = str(sheet.cell(row_idx, 1).value or '').strip()
        item_name = str(sheet.cell(row_idx, 2).value or '').strip()
        user = str(sheet.cell(row_idx, 5).value or '').strip()
        created_at_raw = sheet.cell(row_idx, 6).value
        update_content = str(sheet.cell(row_idx, 7).value or '').strip()

        if not item_id or item_id == 'Item ID':
            continue

        # Parse date
        created_at = ''
        if isinstance(created_at_raw, datetime):
            created_at = created_at_raw.isoformat()
        elif isinstance(created_at_raw, str):
            try:
                # Format: "19/August/2024  03:08:58 PM"
                dt = datetime.strptime(created_at_raw, '%d/%B/%Y  %I:%M:%S %p')
                created_at = dt.isoformat()
            except:
                created_at = created_at_raw

        if not update_content:
            continue

        comment = {
            'author': user,
            'text': update_content,
            'timestamp': created_at
        }

        if item_id not in updates_by_item_id:
            updates_by_item_id[item_id] = []

        updates_by_item_id[item_id].append(comment)

    return updates_by_item_id

# Parse all update files
updates_dir = Path('~/Downloads/account_24917604_data_1763613686/updates').expanduser()
all_updates = {}

print("Parsing Monday.com updates...")
for filepath in sorted(updates_dir.glob('*_updates.xlsx')):
    try:
        board_updates = parse_updates_file(filepath)
        all_updates.update(board_updates)
        total_comments = sum(len(comments) for comments in board_updates.values())
        print(f"✓ {filepath.name}: {len(board_updates)} tasks with {total_comments} comments")
    except Exception as e:
        print(f"✗ Error parsing {filepath.name}: {e}")

print(f"\nTotal: {len(all_updates)} tasks with comments")
print(f"Total comments: {sum(len(comments) for comments in all_updates.values())}")

# Save to JSON
output_file = '/Users/user/Downloads/bright-forge-portal/monday_updates.json'
with open(output_file, 'w') as f:
    json.dump(all_updates, f, indent=2)

print(f"\n✓ Saved updates to {output_file}")

# Show sample
if all_updates:
    sample_id = list(all_updates.keys())[0]
    print(f"\nSample (Item ID {sample_id}):")
    for comment in all_updates[sample_id][:2]:
        print(f"  - {comment['author']}: {comment['text'][:80]}...")
